from openpyxl import Workbook
from openpyxl import load_workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx")

workbook = load_workbook(filename="hello_world.xlsx")


print(sheet["A1"].value)

